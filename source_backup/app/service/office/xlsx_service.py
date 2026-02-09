import logging
from typing import Dict, List

import gradio as gr
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pydantic import conlist

from app.service.langchain.index import LLM_obj
from app.service.langchain.openai_llm_langchain_service import \
    OpenaiLLMLangchainClass
from app.util.config import Settings


async def translate_sheet_content(ws, prompt: str, struct: str, language_radio: str, llm_obj: OpenaiLLMLangchainClass):
    """
    获取 worksheet 中所有内容并进行翻译。

    :param ws: openpyxl worksheet 对象
    :param prompt: 翻译提示
    :param struct: 输出结构
    :param language_radio: 目标语言
    :param llm_obj: LLM 对象
    """
    content_list: List[Dict] = []

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and not isinstance(cell.value, (int, float)):
                content_list.append({
                    'row': cell.row,
                    'col': cell.column,
                    'content': str(cell.value)
                })

    contents_to_translate = [item['content'] for item in content_list]
    translate_text_dict = await (llm_obj.translates_chain(prompt)).chain_ainvoke(contents_to_translate, struct, language_radio)
    translate_text_arr = translate_text_dict['translate_text']

    for i, item in enumerate(content_list):
        ws.cell(row=item['row'], column=item['col'],
                value=translate_text_arr[i])


async def translate_sheet_name(sheet_name: str, prompt: str, struct: str, language_radio: str, llm_obj: OpenaiLLMLangchainClass) -> str:
    """
    翻译sheet名称。

    :param sheet_name: 原sheet名称
    :param prompt: 翻译提示
    :param struct: 输出结构
    :param language_radio: 目标语言
    :param llm_obj: LLM 对象
    :return: 翻译后的sheet名称
    """
    translate_text_dict = await (llm_obj.translates_chain(prompt)).chain_ainvoke([sheet_name], struct, language_radio)
    return translate_text_dict['translate_text'][0]


def save_translated_data(output_path: str, workbook: Workbook):
    """
    保存翻译后的所有sheet到一个Excel文件。

    :param output_path: 输出文件路径
    :param workbook: 包含所有翻译后sheet的 Workbook 对象
    """
    workbook.save(output_path)


async def launch_translate_xlsx(file_path: str, prompt: str, struct: str, output_name: str, language_radio: str, llm_radio: str, progress=gr.Progress()) -> str:
    """
    处理文件，获取所有 sheet 内容并翻译，然后保存结果。

    :param file_path: 文件路径
    :param prompt: 翻译提示
    :param struct: 输出结构
    :param output_name: 输出文件名
    :param language_radio: 目标语言
    :param llm_radio: 使用的LLM模型
    :param progress: Gradio进度条对象
    :return: 输出文件路径
    """
    llm_obj: OpenaiLLMLangchainClass = LLM_obj(llm_name=llm_radio)
    output_path = Settings.FILE_OUTPUT_PATH+output_name

    try:
        if file_path.endswith('.xlsx'):
            workbook = load_workbook(file_path)
            total_sheets = len(workbook.sheetnames)

            for i, sheet_name in enumerate(workbook.sheetnames):
                progress((i + 1) / total_sheets,
                         f"正在翻译第 {i+1}/{total_sheets} sheet, 名称: {sheet_name}")

                ws = workbook[sheet_name]
                await translate_sheet_content(ws, prompt, struct, language_radio, llm_obj)

                translated_sheet_name = await translate_sheet_name(sheet_name, prompt, struct, language_radio, llm_obj)
                ws.title = translated_sheet_name

            save_translated_data(output_path, workbook)

        elif file_path.endswith('.csv'):
            df = pd.read_csv(file_path, header=None)
            translated_df = await translate_sheet_content(df, prompt, struct, language_radio, llm_obj)

            wb = Workbook()
            ws = wb.active
            ws.title = "Translated_Sheet"

            for r in dataframe_to_rows(translated_df, index=False, header=False):
                ws.append(r)

            save_translated_data(output_path, wb)

        return output_path

    except Exception as e:
        logging.error(f"excel 翻译失败: {e}")
        return None
